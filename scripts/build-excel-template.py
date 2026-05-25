#!/usr/bin/env python3
from __future__ import annotations
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "tmp" / "Simulatore-TPL-Lotti-1-4-template-base.xlsx"
OFFERS_CSV = ROOT / "excel-vba" / "templates" / "offerte-esempio.csv"
GOLDEN_CSV = ROOT / "excel-vba" / "templates" / "golden-cases.csv"

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_header(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)


wb = Workbook()
wb.remove(wb.active)

intro = wb.create_sheet("Istruzioni")
intro["A1"] = "Simulatore gara TPL lotti 1-4 - Template Excel (modalità light)"
intro["A1"].font = Font(size=14, bold=True)
intro["A3"] = "Base tecnica per manutenzione: importa i moduli VBA da excel-vba/src, poi salva come template .xlsm."
intro["A5"] = "Checklist rapida"
intro["A5"].font = Font(bold=True)
steps = [
    "1) Apri questa base in Excel, poi ALT+F11 e importa tutti i file .bas da excel-vba/src/.",
    "2) Verifica i nomi foglio: Offerte, Parametri, Ottimizzazione, Risultati, ConfrontoWeb, LogOttimizzazione.",
    "3) Compila/aggiorna Offerte e Parametri.",
    "4) Salva come excel-vba/templates/Simulatore-TPL-Lotti-1-4-template.xlsm, poi esegui: CheckBeforeRun, SimulaScenario, OttimizzaLottoAttivo, ConfrontoWebGolden.",
]
for i, step in enumerate(steps, start=6):
    intro[f"A{i}"] = step

param = wb.create_sheet("Parametri")
param.append(["Parametro", "Valore", "Note"])
style_header(param, 1, 3)
for row in [
    ["SogliaTecnica", 36, "Soglia tecnica minima per ammissibilità (B2)"],
    ["LottoAttivo", "L1", "Lotto attivo per simulazione/ottimizzazione (B3)"],
    ["PunteggioEconomicoMax", 30, "Max punti economici"],
    ["PunteggioTecnicoMax", 70, "Max punti tecnici"],
]:
    param.append(row)

opt = wb.create_sheet("Ottimizzazione")
opt.append(["Parametro", "Valore", "Descrizione"])
style_header(opt, 1, 3)
for row in [
    ["BidderId", "", "Bidder target ottimizzazione (B2)"],
    ["IterMax", 100, "Numero massimo iterazioni (B3)"],
    ["StepQ", 0.5, "Incremento leva Q (B4)"],
    ["Note", "", "B5 riservata"],
    ["RibassoStep", 0.1, "Step riduzione ribasso (B6)"],
    ["TettoQ", 70, "Massimale Q (B7)"],
    ["StepT", 0.5, "Incremento leva T (B8)"],
    ["TettoT", 70, "Massimale T (B9)"],
]:
    opt.append(row)

offers = wb.create_sheet("Offerte")
with OFFERS_CSV.open(newline="", encoding="utf-8") as fh:
    reader = csv.reader(fh)
    for ridx, row in enumerate(reader, start=1):
        offers.append(row)
        if ridx == 1:
            style_header(offers, 1, len(row))

offers.freeze_panes = "A2"

results = wb.create_sheet("Risultati")
results.append(["Lotto", "BidderId", "BidderNome", "Tecnico", "Economico", "Totale", "Ammissibile", "Esito"])
style_header(results, 1, 8)
for lot in ["L1", "L2", "L3", "L4"]:
    results.append([lot, "", "", "", "", "", "", ""])

combo_start = 8
results[f"A{combo_start}"] = "Combinatorie principali"
results[f"A{combo_start}"].font = Font(bold=True)
results.append(["Combinazione", "Ammissibile", "Note"])
style_header(results, combo_start + 1, 3)
for combo in ["L1+L2", "L2+L3", "L3+L4", "L1+L4"]:
    results.append([combo, "", ""])

golden = wb.create_sheet("ConfrontoWeb")
with GOLDEN_CSV.open(newline="", encoding="utf-8") as fh:
    reader = csv.reader(fh)
    for ridx, row in enumerate(reader, start=1):
        golden.append(row)
        if ridx == 1:
            style_header(golden, 1, len(row))

golden.freeze_panes = "A2"

log = wb.create_sheet("LogOttimizzazione")
log.append(["Iterazione", "Lotto", "Leva", "Delta", "TotalePrima", "TotaleDopo", "Esito"])
style_header(log, 1, 7)

for ws in [intro, param, opt, offers, results, golden, log]:
    autofit(ws)

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Creato template: {OUT}")

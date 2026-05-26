#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
import csv
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = ROOT / "excel-vba" / "templates" / "Simulatore-TPL-Lotti-1-4-template.xlsm"
EXCEL_README = ROOT / "excel-vba" / "README.md"
CRITERIA_CSV = ROOT / "excel-vba" / "templates" / "criteria.csv"

# Palette derivata dai token CSS della web app GLM.
NAVY = "172124"
BLUE = "315A8A"
TEAL = "0F766E"
TEAL_DARK = "183B38"
TEAL_SOFT = "E6F4F1"
SURFACE_SOFT = "F8FAF9"
AMBER = "FFF6DF"
AMBER_TEXT = "80510F"
RED = "FFF0ED"
GREEN = TEAL_SOFT
INK = "172124"
MUTED = "657176"
GRID = "DCE5E4"
WHITE = "FFFFFF"
INPUT = "FFF6DF"
OUTPUT = "E6F4F1"
SECTION = "EDF3F2"

MAX_OFFER_ROWS = 200
MAX_COMBO_ROWS = 80
TECHNICAL_START_ROW = 5
AMBIT_MAX_POINTS = {
    "A": 7,
    "B": 14,
    "C": 14,
    "D": 10,
    "E": 4,
    "F": 14,
    "G": 7,
}

THIN_GRID = Side(style="thin", color=GRID)
BORDER = Border(left=THIN_GRID, right=THIN_GRID, top=THIN_GRID, bottom=THIN_GRID)
def workbook_path() -> Path:
    if len(sys.argv) > 1:
        return Path(sys.argv[1]).resolve()
    return DEFAULT_WORKBOOK


def fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def reset_sheet(wb, name: str):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def move_sheet_first(wb, sheet):
    wb._sheets.remove(sheet)
    wb._sheets.insert(0, sheet)


def reorder_sheets(wb, ordered_names: list[str]):
    by_name = {sheet.title: sheet for sheet in wb.worksheets}
    ordered = [by_name[name] for name in ordered_names if name in by_name]
    ordered.extend(sheet for sheet in wb.worksheets if sheet.title not in ordered_names)
    wb._sheets = ordered


def clear_layout_helpers(ws):
    ws.tables.clear()
    ws.data_validations.dataValidation = []
    ws.conditional_formatting._cf_rules.clear()
    ws.auto_filter.ref = None


def load_criteria() -> list[dict[str, str]]:
    with CRITERIA_CSV.open(newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def title(ws, text: str, subtitle: str | None = None, end_col: str = "H"):
    ws["A1"] = text
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = fill(TEAL_DARK)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells(f"A1:{end_col}1")
    ws.row_dimensions[1].height = 28

    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=11, color=MUTED)
        ws["A2"].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"A2:{end_col}2")


def style_cells(ws, cell_range: str, fill_color: str | None = None, bold: bool = False, font_color: str = INK):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = BORDER
            cell.font = Font(bold=bold, color=font_color)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill_color:
                cell.fill = fill(fill_color)


def section_header(ws, row: int, text: str, end_col: str = "H"):
    ws[f"A{row}"] = text
    ws[f"A{row}"].font = Font(size=12, bold=True, color=WHITE)
    ws[f"A{row}"].fill = fill(TEAL)
    ws[f"A{row}"].alignment = Alignment(vertical="center")
    ws.merge_cells(f"A{row}:{end_col}{row}")
    ws.row_dimensions[row].height = 22


def set_widths(ws, widths: dict[str, float]):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def coerce_number(value):
    if value is None or value == "":
        return value
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace("%", "")
    if not text:
        return value
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return value


def add_sheet_link(cell, label: str, sheet_name: str):
    cell.value = label
    cell.hyperlink = f"#'{sheet_name}'!A1"
    cell.style = "Hyperlink"
    cell.font = Font(color=BLUE, underline="single", bold=True)


def apply_sheet_chrome(ws, tab_color: str = TEAL):
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and not cell.font.color:
                cell.font = Font(color=INK)


def add_header_comment(ws, ref: str, text: str):
    ws[ref].comment = Comment(text, "GLM")


def hide_columns(ws, columns: list[str]):
    for column in columns:
        ws.column_dimensions[column].hidden = True


def apply_workbook_visibility(wb):
    visible = {
        "Dashboard",
        "Parametri",
        "Offerte",
        "CriteriTecnici",
        "Ottimizzazione",
        "Combinatorie",
        "Risultati",
        "Guida",
    }
    for ws in wb.worksheets:
        ws.sheet_state = "visible" if ws.title in visible else "hidden"


def create_dashboard(wb):
    ws = reset_sheet(wb, "Dashboard")
    move_sheet_first(wb, ws)
    apply_sheet_chrome(ws, TEAL_DARK)
    title(
        ws,
        "Simulatore gara TPL lotti 1-4",
        "Workbook operativo allineato alla web app: parti da qui, poi entra solo nei fogli necessari.",
    )

    section_header(ws, 4, "Stato scenario", "H")
    cards = [
        ("A5:B7", "Scenario", '=IF(Parametri!B3="","Da impostare","Lotto "&Parametri!B3)', "Lotto attivo per analisi e ottimizzazione"),
        ("C5:D7", "Offerte", "=COUNTA(Offerte!A2:A200)", "Righe concorrente-lotto compilate"),
        ("E5:F7", "Attive", '=COUNTIF(Offerte!D2:D200,"1")+COUNTIF(Offerte!D2:D200,1)', "Righe incluse nel calcolo"),
        ("G5:H7", "Soglia", '=Parametri!B2&" punti"', "Minimo tecnico di ammissibilità"),
    ]
    for cell_range, label, formula, note in cards:
        start_ref, end_ref = cell_range.split(":")
        start = ws[start_ref]
        end = ws[end_ref]
        for row in range(start.row, end.row + 1):
            ws.merge_cells(start_row=row, start_column=start.column, end_row=row, end_column=end.column)
            cell = ws.cell(row=row, column=start.column)
            cell.fill = fill(TEAL_SOFT)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=start.row, column=start.column).value = label
        ws.cell(row=start.row, column=start.column).font = Font(color=INK, bold=True, size=10)
        ws.cell(row=start.row + 1, column=start.column).value = formula
        ws.cell(row=start.row + 1, column=start.column).font = Font(color=TEAL_DARK, bold=True, size=16)
        ws.cell(row=start.row + 2, column=start.column).value = note
        ws.cell(row=start.row + 2, column=start.column).font = Font(color=MUTED, size=9)
        for row in range(start.row, end.row + 1):
            ws.row_dimensions[row].height = 24

    section_header(ws, 10, "Organizzazione come nella web app", "H")
    workflow = [
        ("1", "Scenario", "Parametri", "Soglia tecnica, lotto attivo e impostazioni di base."),
        ("2", "Tecnica", "CriteriTecnici", "Sub-criteri A-G, con input evidenziati e punteggio alimentato in Offerte."),
        ("3", "Economica", "Offerte", "Concorrenti, lotti, attivazione righe e ribasso medio."),
        ("4", "Ottimizzazione", "Ottimizzazione", "Bidder target, leve Q/T e iterazioni controllate."),
        ("5", "Combinatorie", "Combinatorie", "Coppie principali, buste, PEF e ribasso migliorativo."),
        ("6", "Risultati", "Risultati", "Ranking generato dalla macro SimulaScenario."),
    ]
    start = 11
    for idx, (step, area, sheet_name, detail) in enumerate(workflow, start=start):
        ws[f"A{idx}"] = step
        ws[f"B{idx}"] = area
        add_sheet_link(ws[f"C{idx}"], sheet_name, sheet_name)
        ws[f"D{idx}"] = detail
        ws.merge_cells(f"D{idx}:H{idx}")
    style_cells(ws, f"A{start}:H{start + len(workflow) - 1}", WHITE)
    for row in range(start, start + len(workflow)):
        ws[f"A{row}"].fill = fill(TEAL_DARK)
        ws[f"A{row}"].font = Font(color=WHITE, bold=True)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].fill = fill(SECTION)
        ws[f"B{row}"].font = Font(bold=True, color=INK)

    section_header(ws, 20, "Azioni principali", "H")
    links = [
        ("A21", "Parametri", "Parametri"),
        ("C21", "Offerte", "Offerte"),
        ("E21", "Criteri tecnici", "CriteriTecnici"),
        ("G21", "Risultati", "Risultati"),
        ("A23", "Ottimizzazione", "Ottimizzazione"),
        ("C23", "Combinatorie", "Combinatorie"),
        ("E23", "Guida", "Guida"),
    ]
    for ref, label, sheet_name in links:
        add_sheet_link(ws[ref], label, sheet_name)
    macro_rows = [
        ("CheckBeforeRun", "Controlla setup e dati prima di simulare."),
        ("SimulaScenario", "Calcola punteggi, combinatorie e vincitori."),
        ("OttimizzaLottoAttivo", "Applica iterazioni Q/T al bidder indicato."),
        ("ConfrontoWebGolden", "Verifica i totali attesi copiati dal web."),
    ]
    for row, (macro, description) in enumerate(macro_rows, start=25):
        ws[f"A{row}"] = macro
        ws[f"C{row}"] = description
        ws.merge_cells(f"C{row}:H{row}")
    style_cells(ws, "A25:H28", WHITE)
    for row in range(25, 29):
        ws[f"A{row}"].fill = fill(GREEN)
        ws[f"A{row}"].font = Font(bold=True, color=INK)

    section_header(ws, 31, "Fogli avanzati nascosti", "H")
    advanced = [
        "I fogli ScambioWeb, ScenarioGlobale, ConfrontoWeb, LogOttimizzazione, Glossario e Istruzioni restano nel file, ma sono nascosti per non appesantire il lavoro quotidiano.",
        "Usali solo per audit, export JSON verso web, confronto golden, log macro o manutenzione. Puoi riattivarli da Excel con Mostra foglio.",
    ]
    for row, text in enumerate(advanced, start=32):
        ws[f"A{row}"] = text
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"].fill = fill(SURFACE_SOFT if row == 32 else WHITE)
        ws[f"A{row}"].border = BORDER
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)

    section_header(ws, 35, "Limiti da ricordare", "H")
    limits = [
        "Il foglio CriteriTecnici calcola il tecnico dai sub-criteri A-G; il valore aggregato resta solo come fallback di compatibilità.",
        "Il foglio ScambioWeb produce un JSON completo con offerte, sub-criteri, ribassi e combinatorie, ma resta nascosto finché non serve.",
        "I costi e le leve sono ipotesi operative: non sono dati ufficiali di gara.",
        "Se Excel blocca le macro dopo il download, sblocca il file dalle proprietà del sistema prima dell'uso.",
    ]
    for row, text in enumerate(limits, start=36):
        ws[f"A{row}"] = text
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"].fill = fill(AMBER if row == 36 else WHITE)
        ws[f"A{row}"].border = BORDER
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)

    set_widths(ws, {"A": 14, "B": 22, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18})
    ws.freeze_panes = "A4"


def create_guide(wb):
    ws = reset_sheet(wb, "Guida")
    apply_sheet_chrome(ws, BLUE)
    title(ws, "Guida operativa", "Percorso d'uso, mappa dei fogli e README incorporato nel workbook.")

    rows = [
        ("Avvio rapido", "1. Apri Dashboard e controlla lo stato scenario.\n2. In Parametri imposta soglia e lotto attivo.\n3. In Offerte compila concorrenti, lotti, attivo e ribasso.\n4. In CriteriTecnici compila solo le celle evidenziate in giallo.\n5. Esegui CheckBeforeRun e poi SimulaScenario.\n6. Leggi Risultati."),
        ("Mappa rispetto alla web app", "Scenario = Parametri.\nTecnica = CriteriTecnici.\nEconomica = Offerte.\nOttimizzazione = Ottimizzazione.\nCombinatorie = Combinatorie.\nRisultati = Risultati."),
        ("Cosa non devi toccare", "Le colonne nascoste e i fogli nascosti servono a formule, macro, scambio dati e audit. Restano nel file per completezza, ma non fanno parte del percorso quotidiano."),
        ("Macro principali", "CheckBeforeRun: valida setup e input.\nSimulaScenario: calcola risultati.\nOttimizzaLottoAttivo: lavora sul bidder e lotto selezionati.\nConfrontoWebGolden: confronta i totali Excel con valori web incollati nel foglio nascosto ConfrontoWeb."),
        ("Sicurezza macro", "Dopo download da web Excel può bloccare le macro. Su macOS/Windows può servire sbloccare il file o spostarlo in una posizione attendibile."),
    ]
    row = 4
    for heading, body in rows:
        section_header(ws, row, heading, "H")
        row += 1
        ws[f"A{row}"] = body
        ws.merge_cells(f"A{row}:H{row + 2}")
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"A{row}"].border = BORDER
        row += 4

    section_header(ws, row, "README incorporato", "H")
    row += 1
    for line in EXCEL_README.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        ws[f"A{row}"] = line
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        if line.startswith("#"):
            ws[f"A{row}"].font = Font(bold=True, color=NAVY, size=12)
            ws[f"A{row}"].fill = fill(SECTION)
        else:
            ws[f"A{row}"].font = Font(color=INK, size=10)
        row += 1

    set_widths(ws, {"A": 18, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16})
    ws.freeze_panes = "A4"


def create_glossary(wb):
    ws = reset_sheet(wb, "Glossario")
    apply_sheet_chrome(ws, TEAL)
    title(ws, "Glossario e campi", "Riferimento rapido per compilare senza cercare istruzioni esterne.")
    headers = ["Area", "Campo/Macro", "Dove", "Significato", "Note operative"]
    ws.append([])
    ws.append(headers)
    rows = [
        ("Parametri", "SogliaTecnica", "Parametri!B2", "Punteggio tecnico minimo per essere ammessi.", "Valore 0-70."),
        ("Parametri", "LottoAttivo", "Parametri!B3", "Lotto usato dalle macro di ottimizzazione.", "Dropdown L1-L4."),
        ("Offerte", "Attivo", "Offerte!D:D", "1 include la riga nella simulazione, 0 la esclude.", "Usare 0/1."),
        ("Offerte", "TecnicoRiparametrato", "Offerte!E:E", "Punteggio tecnico finale alimentato dai sub-criteri A-G.", "Usa fallback aggregato solo se i criteri sono vuoti."),
        ("Offerte", "TecnicoSogliaQT", "Offerte!O:O", "Somma Q/T usata per la soglia di sbarramento.", "Nascosto: i discrezionali D non entrano nella soglia."),
        ("Offerte", "RibassoMedioPercento", "Offerte!F:F", "Ribasso economico medio simulato.", "0-100."),
        ("Offerte", "TotaleFormula", "Offerte!J:J", "Totale calcolato direttamente nel foglio.", "Tecnico + economico."),
        ("CriteriTecnici", "Sub-criteri A-G", "CriteriTecnici!A:S", "Input e punteggi tecnici per ciascuna offerta.", "Q, T e D seguono le formule del simulatore."),
        ("Combinatorie", "Attivo", "Combinatorie!D:D", "1 include la coppia nella matrice scenario.", "Richiede lotti singoli ammessi."),
        ("Combinatorie", "RibassoCombinatoria", "Combinatorie!E:E", "Ribasso medio della coppia.", "Deve migliorare il riferimento singolo indicativo."),
        ("Scenario globale", "Matrice scenari", "ScenarioGlobale", "Confronta singoli e combinatorie compatibili.", "Indicativa: per vincoli avanzati resta centrale la web app."),
        ("Scambio web", "JSON Excel", "ScambioWeb!A:A", "Payload copiabile per import/export con il simulatore web.", "Include offerte, criteri A-G e combinatorie."),
        ("Ottimizzazione", "BidderId", "Ottimizzazione!B2", "Concorrente target.", "Deve esistere in Offerte."),
        ("Macro", "CheckBeforeRun", "Macro", "Controllo input e struttura workbook.", "Eseguire prima di simulare."),
        ("Macro", "SimulaScenario", "Macro", "Calcola risultati, combinatorie e vincitori.", "Aggiorna Risultati."),
        ("Macro", "OttimizzaLottoAttivo", "Macro", "Itera sulle leve Q/T del bidder target.", "Aggiorna Offerte e Log."),
        ("Macro", "ConfrontoWebGolden", "Macro", "Confronta Excel con expected web.", "Usa J2:J5 in ConfrontoWeb."),
    ]
    for item in rows:
        ws.append(item)
    style_header_row(ws, 3, len(headers))
    style_cells(ws, f"A4:E{3 + len(rows)}", WHITE)
    set_widths(ws, {"A": 18, "B": 24, "C": 24, "D": 45, "E": 42})
    ws.freeze_panes = "A4"


def style_header_row(ws, row: int, cols: int):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill(NAVY)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 24


def add_table(ws, name: str, ref: str):
    if name in ws.tables:
        del ws.tables[name]
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def polish_instruction_sheet(wb):
    ws = wb["Istruzioni"]
    apply_sheet_chrome(ws, BLUE)
    ws.delete_rows(1, ws.max_row)
    title(ws, "Istruzioni rapide", "Punto di partenza per usare il file senza leggere documenti esterni.")
    rows = [
        ("1", "Dashboard", "Apri la dashboard per orientarti e passare ai fogli principali."),
        ("2", "Parametri", "Imposta soglia tecnica e lotto attivo."),
        ("3", "Offerte", "Compila concorrenti, lotti attivi, tecnico e ribasso."),
        ("4", "Macro", "Esegui CheckBeforeRun, SimulaScenario e, se serve, OttimizzaLottoAttivo."),
        ("5", "Confronto web", "Incolla i valori attesi in ConfrontoWeb!J2:J5 e lancia ConfrontoWebGolden."),
    ]
    section_header(ws, 4, "Percorso consigliato", "H")
    for row, item in enumerate(rows, start=5):
        ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"] = item
        ws.merge_cells(f"C{row}:H{row}")
    style_cells(ws, "A5:H9", WHITE)
    for row in range(5, 10):
        ws[f"A{row}"].fill = fill(NAVY)
        ws[f"A{row}"].font = Font(color=WHITE, bold=True)
        ws[f"B{row}"].fill = fill(SECTION)
        ws[f"B{row}"].font = Font(bold=True)
    section_header(ws, 12, "Avviso", "H")
    ws["A13"] = "Il workbook supporta simulazione offline con offerte, sub-criteri A-G e combinatorie. Per warning documentali avanzati, persistenza e confronto scenari salvati resta centrale la web app."
    ws.merge_cells("A13:H14")
    ws["A13"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A13"].fill = fill(AMBER)
    ws["A13"].border = BORDER
    set_widths(ws, {"A": 8, "B": 20, "C": 24, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16})
    ws.freeze_panes = "A4"


def polish_parametri(wb):
    ws = wb["Parametri"]
    clear_layout_helpers(ws)
    apply_sheet_chrome(ws, AMBER_TEXT)
    style_header_row(ws, 1, 3)
    style_cells(ws, f"A2:C{ws.max_row}", WHITE)
    for row in range(2, ws.max_row + 1):
        ws[f"B{row}"].fill = fill(INPUT)
        ws[f"B{row}"].font = Font(bold=True, color=INK)
    dv_lot = DataValidation(type="list", formula1='"L1,L2,L3,L4"', allow_blank=False)
    ws.add_data_validation(dv_lot)
    dv_lot.add(ws["B3"])
    dv_threshold = DataValidation(type="decimal", operator="between", formula1="0", formula2="70", allow_blank=False)
    ws.add_data_validation(dv_threshold)
    dv_threshold.add(ws["B2"])
    ws["B2"].comment = Comment("Valore fra 0 e 70. Usato da tutte le macro di ammissibilità.", "Codex")
    ws["B3"].comment = Comment("Lotto attivo per OttimizzaLottoAttivo.", "Codex")
    add_table(ws, "tblParametri", f"A1:C{ws.max_row}")
    set_widths(ws, {"A": 24, "B": 16, "C": 60})
    ws.freeze_panes = "A2"


def polish_ottimizzazione(wb):
    ws = wb["Ottimizzazione"]
    clear_layout_helpers(ws)
    apply_sheet_chrome(ws, TEAL)
    style_header_row(ws, 1, 3)
    style_cells(ws, f"A2:C{ws.max_row}", WHITE)
    for row in [2, 3, 4, 6, 7, 8, 9]:
        ws[f"B{row}"].fill = fill(INPUT)
        ws[f"B{row}"].font = Font(bold=True, color=INK)
    dv_iter = DataValidation(type="whole", operator="between", formula1="1", formula2="1000")
    dv_decimal = DataValidation(type="decimal", operator="between", formula1="0", formula2="100")
    ws.add_data_validation(dv_iter)
    ws.add_data_validation(dv_decimal)
    dv_iter.add(ws["B3"])
    for ref in ["B4", "B6", "B7", "B8", "B9"]:
        dv_decimal.add(ws[ref])
    add_table(ws, "tblOttimizzazione", f"A1:C{ws.max_row}")
    set_widths(ws, {"A": 24, "B": 16, "C": 64})
    ws.freeze_panes = "A2"


def polish_offerte(wb):
    ws = wb["Offerte"]
    clear_layout_helpers(ws)
    apply_sheet_chrome(ws, BLUE)
    headers = [
        "BidderId",
        "BidderNome",
        "Lotto",
        "Attivo",
        "TecnicoRiparametrato",
        "RibassoMedioPercento",
        "AmmessoFormula",
        "RMaxAmmessi",
        "EconomicoFormula",
        "TotaleFormula",
        "Warning",
        "ChiaveOfferta",
        "ChiaveLottoTotale",
        "TecnicoAggregatoFallback",
        "TecnicoSogliaQT",
        "FonteTecnico",
    ]
    fallback_values = {}
    for row in range(2, MAX_OFFER_ROWS + 1):
        fallback_source = ws[f"E{row}"].value
        if isinstance(fallback_source, str) and fallback_source.startswith("="):
            fallback_source = ws[f"N{row}"].value
        fallback_values[row] = coerce_number(fallback_source)
    for row in range(2, MAX_OFFER_ROWS + 1):
        for col in ["D", "E", "F"]:
            ws[f"{col}{row}"].value = coerce_number(ws[f"{col}{row}"].value)
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header
    style_header_row(ws, 1, len(headers))
    add_header_comment(ws, "A1", "Identificativo breve del concorrente. Usalo uguale in tutti i lotti dello stesso operatore.")
    add_header_comment(ws, "B1", "Nome leggibile del concorrente.")
    add_header_comment(ws, "C1", "Lotto L1-L4.")
    add_header_comment(ws, "D1", "1 = riga inclusa nella simulazione, 0 = esclusa.")
    add_header_comment(ws, "E1", "Punteggio tecnico finale calcolato dai sub-criteri A-G. Non modificarlo se usi CriteriTecnici.")
    add_header_comment(ws, "F1", "Ribasso economico medio in percentuale.")
    add_header_comment(ws, "G1", "Esito della soglia tecnica.")
    add_header_comment(ws, "J1", "Totale tecnico + economico calcolato nel foglio.")
    style_cells(ws, f"A2:P{MAX_OFFER_ROWS}", WHITE)

    criteria_end_row = TECHNICAL_START_ROW + (MAX_OFFER_ROWS - 1) * len(load_criteria()) - 1
    criteria_score_sum = 'SUMIFS(CriteriTecnici!$O:$O,CriteriTecnici!$Q:$Q,$A{row}&"|"&$C{row})'
    criteria_qt_sum = (
        'SUMIFS(CriteriTecnici!$O:$O,CriteriTecnici!$Q:$Q,$A{row}&"|"&$C{row},CriteriTecnici!$F:$F,"Q")+'
        'SUMIFS(CriteriTecnici!$O:$O,CriteriTecnici!$Q:$Q,$A{row}&"|"&$C{row},CriteriTecnici!$F:$F,"T")'
    )
    technical_ranges = {
        "lot": f"CriteriTecnici!$C${TECHNICAL_START_ROW}:$C${criteria_end_row}",
        "ambit": f"CriteriTecnici!$E${TECHNICAL_START_ROW}:$E${criteria_end_row}",
        "score": f"CriteriTecnici!$O${TECHNICAL_START_ROW}:$O${criteria_end_row}",
        "admitted": f"CriteriTecnici!$T${TECHNICAL_START_ROW}:$T${criteria_end_row}",
    }

    def riparam_expr(row: int, ambit: str, max_points: int) -> str:
        raw = f'SUMIFS(CriteriTecnici!$O:$O,CriteriTecnici!$Q:$Q,$A{row}&"|"&$C{row},CriteriTecnici!$E:$E,"{ambit}")'
        best = (
            f'SUMPRODUCT(MAX(({technical_ranges["lot"]}=$C{row})*({technical_ranges["ambit"]}="{ambit}")*'
            f'({technical_ranges["admitted"]}="SI")*IFERROR({technical_ranges["score"]}*1,0)))'
        )
        return f'IF({best}>0,MIN({max_points},{raw}*({max_points}/{best})),0)'

    for row in range(2, MAX_OFFER_ROWS + 1):
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = fill(INPUT)
        ws.cell(row=row, column=5).fill = fill(OUTPUT)
        for col in range(7, 17):
            ws.cell(row=row, column=col).fill = fill(OUTPUT)

        criteria_total = criteria_score_sum.format(row=row)
        criteria_qt = criteria_qt_sum.format(row=row)
        riparam_total = "+".join(riparam_expr(row, ambit, max_points) for ambit, max_points in AMBIT_MAX_POINTS.items())
        ws[f"E{row}"] = (
            f'=IF($A{row}="",0,IFERROR(IF({criteria_total}>0,'
            f'IF($G{row}="SI",ROUND({riparam_total},4),0),IFERROR($N{row}*1,0)),0))'
        )
        ws[f"G{row}"] = f'=IF($A{row}="","",IF(IFERROR(VALUE($D{row}),0)<>1,"NO",IF(IFERROR(VALUE($O{row}),0)>=Parametri!$B$2,"SI","NO")))'
        ws[f"H{row}"] = (
            f'=IF($G{row}="SI",'
            f'SUMPRODUCT(MAX(($C$2:$C${MAX_OFFER_ROWS}=$C{row})*(IFERROR($D$2:$D${MAX_OFFER_ROWS}*1,0)=1)*'
            f'(IFERROR($O$2:$O${MAX_OFFER_ROWS}*1,0)>=Parametri!$B$2)*IFERROR($F$2:$F${MAX_OFFER_ROWS}*1,0))),0)'
        )
        ws[f"I{row}"] = f'=IF($G{row}="SI",IF($H{row}>0,30*(IFERROR($F{row}*1,0)/$H{row}),0),0)'
        ws[f"J{row}"] = f'=IF($A{row}="",0,IF($G{row}="SI",ROUND(IFERROR($E{row}*1,0)+$I{row},4),ROUND(IFERROR($E{row}*1,0),4)))'
        ws[f"K{row}"] = (
            f'=IF($A{row}="","",'
            f'IF(IFERROR(VALUE($D{row}),0)<>1,"Non attiva",IF(IFERROR(VALUE($O{row}),0)<Parametri!$B$2,"Sotto soglia","")))'
        )
        ws[f"L{row}"] = f'=IF($A{row}="","",$A{row}&"|"&$C{row})'
        ws[f"M{row}"] = f'=IF($A{row}="","",$C{row}&"|"&TEXT($J{row},"0.0000"))'
        if fallback_values.get(row) not in (None, "") and not str(fallback_values[row]).startswith("="):
            ws[f"N{row}"] = fallback_values[row]
        ws[f"O{row}"] = (
            f'=IF($A{row}="",0,IFERROR(IF({criteria_total}>0,{criteria_qt},IFERROR($N{row}*1,0)),0))'
        )
        ws[f"P{row}"] = (
            f'=IF($A{row}="","",IF({criteria_total}>0,'
            f'"Criteri A-G","Aggregato"))'
        )
    dv_lot = DataValidation(type="list", formula1='"L1,L2,L3,L4"', allow_blank=True)
    dv_active = DataValidation(type="list", formula1='"1,0"', allow_blank=True)
    dv_tech = DataValidation(type="decimal", operator="between", formula1="0", formula2="70", allow_blank=True)
    dv_discount = DataValidation(type="decimal", operator="between", formula1="0", formula2="100", allow_blank=True)
    for dv in [dv_lot, dv_active, dv_tech, dv_discount]:
        ws.add_data_validation(dv)
    dv_lot.add(f"C2:C{MAX_OFFER_ROWS}")
    dv_active.add(f"D2:D{MAX_OFFER_ROWS}")
    dv_tech.add(f"N2:N{MAX_OFFER_ROWS}")
    dv_discount.add(f"F2:F{MAX_OFFER_ROWS}")
    ws.conditional_formatting.add(f"D2:D{MAX_OFFER_ROWS}", CellIsRule(operator="equal", formula=['"0"'], fill=fill(RED)))
    ws.conditional_formatting.add(f"O2:O{MAX_OFFER_ROWS}", FormulaRule(formula=['AND($O2<Parametri!$B$2,$A2<>"")'], fill=fill(AMBER)))
    ws.conditional_formatting.add(f"G2:G{MAX_OFFER_ROWS}", CellIsRule(operator="equal", formula=['"SI"'], fill=fill(GREEN)))
    ws.conditional_formatting.add(f"G2:G{MAX_OFFER_ROWS}", CellIsRule(operator="equal", formula=['"NO"'], fill=fill(RED)))
    add_table(ws, "tblOfferte", f"A1:P{MAX_OFFER_ROWS}")
    set_widths(ws, {"A": 16, "B": 26, "C": 12, "D": 12, "E": 22, "F": 24, "G": 18, "H": 14, "I": 18, "J": 18, "K": 22, "L": 22, "M": 18, "N": 26, "O": 18, "P": 16})
    hide_columns(ws, ["L", "M", "N", "O", "P"])
    ws.freeze_panes = "A2"


def excel_sumproduct_max(condition_formula: str, value_range: str) -> str:
    return f"SUMPRODUCT(MAX({condition_formula}*IFERROR({value_range}*1,0)))"


def create_criteri_tecnici(wb):
    criteria = load_criteria()
    ws = reset_sheet(wb, "CriteriTecnici")
    apply_sheet_chrome(ws, TEAL)
    title(
        ws,
        "Criteri tecnici A-G",
        "Compila i sub-criteri tecnici per ogni offerta: il totale alimenta automaticamente Offerte!E:E.",
    )

    headers = [
        "BidderId",
        "BidderNome",
        "Lotto",
        "CriterioId",
        "Ambito",
        "Tipo",
        "Formula",
        "PuntiMax",
        "Descrizione",
        "Valore",
        "Numeratore",
        "Denominatore",
        "Flag/Coeff",
        "ValoreCalcolato",
        "PunteggioRaw",
        "Note",
        "ChiaveOfferta",
        "ChiaveCriterio",
        "OffertaAttiva",
        "OffertaAmmessa",
    ]
    header_row = 4
    for col, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col).value = header
    style_header_row(ws, header_row, len(headers))
    add_header_comment(ws, "J4", "Input diretto quando il criterio usa un valore unico.")
    add_header_comment(ws, "K4", "Numeratore per criteri percentuali o rapporto.")
    add_header_comment(ws, "L4", "Denominatore per criteri percentuali o rapporto.")
    add_header_comment(ws, "M4", "Per criteri tabellari: 1 se presente, 0 se assente. Per discrezionali: coefficiente 0-1.")
    add_header_comment(ws, "O4", "Punteggio raw del sub-criterio usato dal totale tecnico.")

    total_rows = (MAX_OFFER_ROWS - 1) * len(criteria)
    end_row = TECHNICAL_START_ROW + total_rows - 1
    style_cells(ws, f"A{TECHNICAL_START_ROW}:T{end_row}", WHITE)

    active_range = f"$S${TECHNICAL_START_ROW}:$S${end_row}"
    lot_range = f"$C${TECHNICAL_START_ROW}:$C${end_row}"
    criterion_range = f"$D${TECHNICAL_START_ROW}:$D${end_row}"
    value_range = f"$N${TECHNICAL_START_ROW}:$N${end_row}"
    key_range = f"$Q${TECHNICAL_START_ROW}:$Q${end_row}"

    for index in range(total_rows):
        row = TECHNICAL_START_ROW + index
        offer_row = 2 + index // len(criteria)
        criterion = criteria[index % len(criteria)]
        criterion_id = criterion["CriterionId"]
        kind = criterion["Kind"]
        formula = criterion["Formula"]
        max_points = float(criterion["MaxPoints"])
        quantity_kind = criterion.get("QuantityKind", "")
        dependency_criterion = criterion.get("DependencyCriterion", "")
        dependency_min = criterion.get("DependencyMin", "")

        ws[f"A{row}"] = f'=IF(Offerte!$A{offer_row}="","",Offerte!$A{offer_row})'
        ws[f"B{row}"] = f'=IF(Offerte!$A{offer_row}="","",Offerte!$B{offer_row})'
        ws[f"C{row}"] = f'=IF(Offerte!$A{offer_row}="","",Offerte!$C{offer_row})'
        ws[f"D{row}"] = criterion_id
        ws[f"E{row}"] = criterion["Ambit"]
        ws[f"F{row}"] = kind
        ws[f"G{row}"] = formula
        ws[f"H{row}"] = max_points
        ws[f"I{row}"] = criterion["Label"]
        ws[f"P{row}"] = (
            f'{criterion.get("NumeratorLabel", "")} / {criterion.get("DenominatorLabel", "")}'
            if quantity_kind
            else criterion.get("Unit", "")
        )
        ws[f"Q{row}"] = f'=IF($A{row}="","",$A{row}&"|"&$C{row})'
        ws[f"R{row}"] = f'=IF($A{row}="","",$A{row}&"|"&$C{row}&"|"&$D{row})'
        ws[f"S{row}"] = f'=IF($A{row}="","",IFERROR(INDEX(Offerte!$D$2:$D${MAX_OFFER_ROWS},MATCH($Q{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),0))'
        ws[f"T{row}"] = f'=IF($A{row}="","",IFERROR(INDEX(Offerte!$G$2:$G${MAX_OFFER_ROWS},MATCH($Q{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),"NO"))'

        if kind == "Q" and quantity_kind == "percent":
            ws[f"N{row}"] = f'=IF($A{row}="","",IFERROR(VALUE($K{row})/VALUE($L{row})*100,IFERROR(VALUE($J{row}),0)))'
        elif kind == "Q" and quantity_kind == "ratio":
            ws[f"N{row}"] = f'=IF($A{row}="","",IFERROR(VALUE($K{row})/VALUE($L{row}),IFERROR(VALUE($J{row}),0)))'
        elif kind == "Q":
            ws[f"N{row}"] = f'=IF($A{row}="","",IFERROR(VALUE($J{row}),0))'
        elif kind == "T":
            ws[f"N{row}"] = f'=IF($A{row}="","",IF(IFERROR(VALUE($M{row}),0)=1,1,0))'
        else:
            ws[f"N{row}"] = f'=IF($A{row}="","",MAX(0,MIN(1,IFERROR(VALUE($M{row}),0))))'

        condition = f"({lot_range}=$C{row})*({criterion_range}=$D{row})*({active_range}=1)"
        max_value = excel_sumproduct_max(condition, value_range)
        no_input = f"COUNTA($J{row}:$M{row})=0"
        if kind == "Q" and formula == "higher":
            ws[f"O{row}"] = f'=IF($A{row}="",0,IF(OR($S{row}<>1,{no_input}),0,IF({max_value}>0,ROUND($H{row}*($N{row}/{max_value}),4),0)))'
        elif kind == "Q" and formula == "lower":
            min_value = (
                f'IFERROR(AGGREGATE(15,6,{value_range}/(({lot_range}=$C{row})*({criterion_range}=$D{row})*'
                f'({active_range}=1)*({value_range}>0)),1),0)'
            )
            ws[f"O{row}"] = f'=IF($A{row}="",0,IF(OR($S{row}<>1,{no_input}),0,IF(AND({min_value}>0,$N{row}>0),ROUND($H{row}*({min_value}/$N{row}),4),0)))'
        elif kind == "Q" and formula == "soil":
            ws[f"O{row}"] = f'=IF($A{row}="",0,IF(OR($S{row}<>1,{no_input}),0,IF($N{row}<=0,$H{row},IF({max_value}>0,ROUND($H{row}*(({max_value}-$N{row})/{max_value}),4),0))))'
        elif kind == "T" and dependency_criterion:
            dependency_value = f'IFERROR(SUMIFS({value_range},{key_range},$Q{row},{criterion_range},"{dependency_criterion}"),0)'
            ws[f"O{row}"] = f'=IF($A{row}="",0,IF(OR($S{row}<>1,{no_input}),0,IF(AND($N{row}=1,{dependency_value}>={dependency_min}),$H{row},0)))'
            ws[f"P{row}"] = f'Dipende da {dependency_criterion} >= {dependency_min}'
        elif kind == "T":
            ws[f"O{row}"] = f'=IF($A{row}="",0,IF(OR($S{row}<>1,{no_input}),0,IF($N{row}=1,$H{row},0)))'
        else:
            max_coeff = excel_sumproduct_max(condition, value_range)
            ws[f"O{row}"] = f'=IF($A{row}="",0,IF(OR($S{row}<>1,{no_input}),0,IF({max_coeff}>0,ROUND($H{row}*($N{row}/{max_coeff}),4),0)))'

        for input_col in ["J", "K", "L", "M"]:
            ws[f"{input_col}{row}"].fill = fill(INPUT)
        for output_col in ["N", "O", "Q", "R", "S", "T"]:
            ws[f"{output_col}{row}"].fill = fill(OUTPUT)

    dv_number = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv_ratio = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    ws.add_data_validation(dv_number)
    ws.add_data_validation(dv_ratio)
    dv_number.add(f"J{TECHNICAL_START_ROW}:L{end_row}")
    dv_ratio.add(f"M{TECHNICAL_START_ROW}:M{end_row}")
    ws.conditional_formatting.add(f"O{TECHNICAL_START_ROW}:O{end_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=fill(GREEN)))
    add_table(ws, "tblCriteriTecnici", f"A{header_row}:T{end_row}")
    set_widths(
        ws,
        {
            "A": 16,
            "B": 24,
            "C": 10,
            "D": 12,
            "E": 10,
            "F": 8,
            "G": 14,
            "H": 10,
            "I": 46,
            "J": 14,
            "K": 16,
            "L": 16,
            "M": 14,
            "N": 18,
            "O": 16,
            "P": 38,
            "Q": 18,
            "R": 24,
            "S": 12,
            "T": 14,
        },
    )
    hide_columns(ws, ["G", "H", "N", "Q", "R", "S", "T"])
    ws.freeze_panes = "A5"


def create_combinatorie(wb):
    ws = reset_sheet(wb, "Combinatorie")
    apply_sheet_chrome(ws, TEAL)
    title(
        ws,
        "Combinatorie",
        "Input per confrontare coppie L1+L2, L2+L3, L3+L4 e L1+L4 nella matrice scenario globale.",
    )
    headers = [
        "BidderId",
        "BidderNome",
        "Coppia",
        "Attivo",
        "RibassoCombinatoria",
        "InseritoBuste",
        "PEFCoerente",
        "LottoA",
        "LottoB",
        "AttivoA",
        "AttivoB",
        "TecnicoA",
        "TecnicoB",
        "Ammissibile",
        "EconomicoA",
        "EconomicoB",
        "TotaleCoppia",
        "Note",
        "RibassoSingoloA",
        "RibassoSingoloB",
        "RibassoMinimoIndicativo",
        "ChiaveCoppiaTotale",
        "ChiaveCoppiaBidder",
    ]
    header_row = 4
    for col, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col).value = header
    style_header_row(ws, header_row, len(headers))
    add_header_comment(ws, "D4", "1 = combinatoria inclusa, 0 = esclusa.")
    add_header_comment(ws, "E4", "Ribasso medio della combinatoria, in percentuale.")
    add_header_comment(ws, "F4", "1 se la coppia è inserita in entrambe le buste.")
    add_header_comment(ws, "G4", "1 se il PEF è coerente.")
    add_header_comment(ws, "N4", "Esito ammissibilità della combinatoria.")
    add_header_comment(ws, "Q4", "Totale combinatoria calcolato.")
    style_cells(ws, f"A{header_row + 1}:W{MAX_COMBO_ROWS}", WHITE)

    pair_a = 'IF($C{row}="L1+L2","L1",IF($C{row}="L2+L3","L2",IF($C{row}="L3+L4","L3",IF($C{row}="L1+L4","L1",""))))'
    pair_b = 'IF($C{row}="L1+L2","L2",IF($C{row}="L2+L3","L3",IF($C{row}="L3+L4","L4",IF($C{row}="L1+L4","L4",""))))'

    offer_rows: dict[tuple[str, str], float] = {}
    bidders: dict[str, str] = {}
    ws_off = wb["Offerte"]
    for offer_row in range(2, MAX_OFFER_ROWS + 1):
        bidder_id = str(ws_off[f"A{offer_row}"].value or "").strip()
        bidder_name = str(ws_off[f"B{offer_row}"].value or "").strip()
        lot_id = str(ws_off[f"C{offer_row}"].value or "").strip()
        if not bidder_id or lot_id not in {"L1", "L2", "L3", "L4"}:
            continue
        bidders.setdefault(bidder_id, bidder_name)
        try:
            offer_rows[(bidder_id, lot_id)] = float(ws_off[f"F{offer_row}"].value or 0)
        except (TypeError, ValueError):
            offer_rows[(bidder_id, lot_id)] = 0

    pairs = [("L1+L2", "L1", "L2"), ("L2+L3", "L2", "L3"), ("L3+L4", "L3", "L4"), ("L1+L4", "L1", "L4")]
    seed_rows: list[tuple[str, str, float]] = []
    for bidder_id in bidders:
        for pair_id, first_lot, second_lot in pairs:
            if (bidder_id, first_lot) in offer_rows and (bidder_id, second_lot) in offer_rows:
                reference = (offer_rows[(bidder_id, first_lot)] + offer_rows[(bidder_id, second_lot)]) / 2
                seed_rows.append((bidder_id, pair_id, round(min(reference + 0.1, 100), 4)))

    for row in range(header_row + 1, MAX_COMBO_ROWS + 1):
        seed_index = row - (header_row + 1)
        if seed_index < len(seed_rows):
            bidder_id, pair_id, discount = seed_rows[seed_index]
            ws[f"A{row}"] = bidder_id
            ws[f"C{row}"] = pair_id
            ws[f"D{row}"] = 0
            ws[f"E{row}"] = discount
            ws[f"F{row}"] = 1
            ws[f"G{row}"] = 1
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = fill(INPUT)
        for col in range(8, 24):
            ws.cell(row=row, column=col).fill = fill(OUTPUT)

        ws[f"B{row}"] = f'=IF($A{row}="","",IFERROR(INDEX(Offerte!$B$2:$B${MAX_OFFER_ROWS},MATCH($A{row},Offerte!$A$2:$A${MAX_OFFER_ROWS},0)),""))'
        ws[f"H{row}"] = "=" + pair_a.format(row=row)
        ws[f"I{row}"] = "=" + pair_b.format(row=row)
        ws[f"J{row}"] = f'=IF($H{row}="","",IFERROR(INDEX(Offerte!$D$2:$D${MAX_OFFER_ROWS},MATCH($A{row}&"|"&$H{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),0))'
        ws[f"K{row}"] = f'=IF($I{row}="","",IFERROR(INDEX(Offerte!$D$2:$D${MAX_OFFER_ROWS},MATCH($A{row}&"|"&$I{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),0))'
        ws[f"L{row}"] = f'=IF($H{row}="","",IFERROR(INDEX(Offerte!$E$2:$E${MAX_OFFER_ROWS},MATCH($A{row}&"|"&$H{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),0))'
        ws[f"M{row}"] = f'=IF($I{row}="","",IFERROR(INDEX(Offerte!$E$2:$E${MAX_OFFER_ROWS},MATCH($A{row}&"|"&$I{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),0))'
        ws[f"S{row}"] = f'=IF($H{row}="","",IFERROR(INDEX(Offerte!$F$2:$F${MAX_OFFER_ROWS},MATCH($A{row}&"|"&$H{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),0))'
        ws[f"T{row}"] = f'=IF($I{row}="","",IFERROR(INDEX(Offerte!$F$2:$F${MAX_OFFER_ROWS},MATCH($A{row}&"|"&$I{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),0))'
        ws[f"U{row}"] = f'=IF($A{row}="","",ROUND((IFERROR(VALUE($S{row}),0)+IFERROR(VALUE($T{row}),0))/2,4))'
        ws[f"N{row}"] = (
            f'=IF($A{row}="","",IF(IFERROR(VALUE($D{row}),0)<>1,"NO",'
            f'IF(AND(IFERROR(VALUE($J{row}),0)=1,IFERROR(VALUE($K{row}),0)=1,'
            f'IFERROR(INDEX(Offerte!$G$2:$G${MAX_OFFER_ROWS},MATCH($A{row}&"|"&$H{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),"NO")="SI",'
            f'IFERROR(INDEX(Offerte!$G$2:$G${MAX_OFFER_ROWS},MATCH($A{row}&"|"&$I{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),"NO")="SI",'
            f'IFERROR(VALUE($F{row}),0)=1,IFERROR(VALUE($G{row}),0)=1,IFERROR(VALUE($E{row}),0)>$U{row}),"SI","NO")))'
        )
        ws[f"O{row}"] = (
            f'=IF($N{row}="SI",IFERROR(30*(IFERROR(VALUE($E{row}),0)/MAX('
            f'SUMPRODUCT(MAX((Offerte!$C$2:$C${MAX_OFFER_ROWS}=$H{row})*(IFERROR(Offerte!$D$2:$D${MAX_OFFER_ROWS}*1,0)=1)*'
            f'(IFERROR(Offerte!$E$2:$E${MAX_OFFER_ROWS}*1,0)>=Parametri!$B$2)*IFERROR(Offerte!$F$2:$F${MAX_OFFER_ROWS}*1,0))),'
            f'SUMPRODUCT(MAX(($H$5:$H${MAX_COMBO_ROWS}=$H{row})*($N$5:$N${MAX_COMBO_ROWS}="SI")*IFERROR($E$5:$E${MAX_COMBO_ROWS}*1,0))),'
            f'SUMPRODUCT(MAX(($I$5:$I${MAX_COMBO_ROWS}=$H{row})*($N$5:$N${MAX_COMBO_ROWS}="SI")*IFERROR($E$5:$E${MAX_COMBO_ROWS}*1,0))))),0),0)'
        )
        ws[f"P{row}"] = (
            f'=IF($N{row}="SI",IFERROR(30*(IFERROR(VALUE($E{row}),0)/MAX('
            f'SUMPRODUCT(MAX((Offerte!$C$2:$C${MAX_OFFER_ROWS}=$I{row})*(IFERROR(Offerte!$D$2:$D${MAX_OFFER_ROWS}*1,0)=1)*'
            f'(IFERROR(Offerte!$E$2:$E${MAX_OFFER_ROWS}*1,0)>=Parametri!$B$2)*IFERROR(Offerte!$F$2:$F${MAX_OFFER_ROWS}*1,0))),'
            f'SUMPRODUCT(MAX(($H$5:$H${MAX_COMBO_ROWS}=$I{row})*($N$5:$N${MAX_COMBO_ROWS}="SI")*IFERROR($E$5:$E${MAX_COMBO_ROWS}*1,0))),'
            f'SUMPRODUCT(MAX(($I$5:$I${MAX_COMBO_ROWS}=$I{row})*($N$5:$N${MAX_COMBO_ROWS}="SI")*IFERROR($E$5:$E${MAX_COMBO_ROWS}*1,0))))),0),0)'
        )
        ws[f"Q{row}"] = f'=IF($N{row}="SI",ROUND($L{row}+$M{row}+$O{row}+$P{row},4),0)'
        ws[f"R{row}"] = (
            f'=IF($A{row}="","",IF(IFERROR(VALUE($D{row}),0)<>1,"Non attiva",'
            f'IF(OR(IFERROR(VALUE($J{row}),0)<>1,IFERROR(VALUE($K{row}),0)<>1),"Singoli non attivi",'
            f'IF(OR(IFERROR(INDEX(Offerte!$G$2:$G${MAX_OFFER_ROWS},MATCH($A{row}&"|"&$H{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),"NO")<>"SI",'
            f'IFERROR(INDEX(Offerte!$G$2:$G${MAX_OFFER_ROWS},MATCH($A{row}&"|"&$I{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),"NO")<>"SI"),"Soglia non superata",'
            f'IF(OR(IFERROR(VALUE($F{row}),0)<>1,IFERROR(VALUE($G{row}),0)<>1),"Buste/PEF non confermati",'
            f'IF(IFERROR(VALUE($E{row}),0)<=$U{row},"Ribasso non migliorativo",""))))))'
        )
        ws[f"V{row}"] = f'=IF($A{row}="","",$C{row}&"|"&TEXT($Q{row},"0.0000"))'
        ws[f"W{row}"] = f'=IF($A{row}="","",$C{row}&"|"&$A{row})'

    dv_pair = DataValidation(type="list", formula1='"L1+L2,L2+L3,L3+L4,L1+L4"', allow_blank=True)
    dv_binary = DataValidation(type="list", formula1='"1,0"', allow_blank=True)
    dv_discount = DataValidation(type="decimal", operator="between", formula1="0", formula2="100", allow_blank=True)
    for dv in [dv_pair, dv_binary, dv_discount]:
        ws.add_data_validation(dv)
    dv_pair.add(f"C5:C{MAX_COMBO_ROWS}")
    for ref in [f"D5:D{MAX_COMBO_ROWS}", f"F5:F{MAX_COMBO_ROWS}", f"G5:G{MAX_COMBO_ROWS}"]:
        dv_binary.add(ref)
    dv_discount.add(f"E5:E{MAX_COMBO_ROWS}")
    ws.conditional_formatting.add(f"N5:N{MAX_COMBO_ROWS}", CellIsRule(operator="equal", formula=['"SI"'], fill=fill(GREEN)))
    ws.conditional_formatting.add(f"N5:N{MAX_COMBO_ROWS}", CellIsRule(operator="equal", formula=['"NO"'], fill=fill(RED)))
    add_table(ws, "tblCombinatorie", f"A4:W{MAX_COMBO_ROWS}")
    set_widths(
        ws,
        {
            "A": 16,
            "B": 26,
            "C": 14,
            "D": 10,
            "E": 22,
            "F": 14,
            "G": 14,
            "H": 10,
            "I": 10,
            "J": 10,
            "K": 10,
            "L": 12,
            "M": 12,
            "N": 14,
            "O": 14,
            "P": 14,
            "Q": 16,
            "R": 32,
            "S": 16,
            "T": 16,
            "U": 22,
            "V": 18,
            "W": 18,
        },
    )
    hide_columns(ws, ["H", "I", "J", "K", "L", "M", "O", "P", "S", "T", "U", "V", "W"])
    ws.freeze_panes = "A5"


def create_scenario_globale(wb):
    ws = reset_sheet(wb, "ScenarioGlobale")
    apply_sheet_chrome(ws, BLUE)
    title(
        ws,
        "Scenario globale",
        "Matrice operativa per confrontare singoli e combinatorie compatibili dentro Excel.",
    )

    section_header(ws, 4, "Migliori singoli per lotto", "H")
    single_headers = ["Lotto", "BidderId", "BidderNome", "Totale", "Tecnico", "Economico", "Warning", "Origine"]
    for col, header in enumerate(single_headers, start=1):
        ws.cell(row=5, column=col).value = header
    style_header_row(ws, 5, len(single_headers))
    for row, lot_id in enumerate(["L1", "L2", "L3", "L4"], start=6):
        ws[f"A{row}"] = lot_id
        ws[f"D{row}"] = f'=SUMPRODUCT(MAX((Offerte!$C$2:$C${MAX_OFFER_ROWS}=$A{row})*(Offerte!$G$2:$G${MAX_OFFER_ROWS}="SI")*Offerte!$J$2:$J${MAX_OFFER_ROWS}))'
        ws[f"B{row}"] = f'=IF($D{row}=0,"",IFERROR(INDEX(Offerte!$A$2:$A${MAX_OFFER_ROWS},MATCH($A{row}&"|"&TEXT($D{row},"0.0000"),Offerte!$M$2:$M${MAX_OFFER_ROWS},0)),""))'
        ws[f"C{row}"] = f'=IF($B{row}="","",IFERROR(INDEX(Offerte!$B$2:$B${MAX_OFFER_ROWS},MATCH($B{row}&"|"&$A{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),""))'
        ws[f"E{row}"] = f'=IF($B{row}="","",IFERROR(INDEX(Offerte!$E$2:$E${MAX_OFFER_ROWS},MATCH($B{row}&"|"&$A{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),0))'
        ws[f"F{row}"] = f'=IF($B{row}="","",IFERROR(INDEX(Offerte!$I$2:$I${MAX_OFFER_ROWS},MATCH($B{row}&"|"&$A{row},Offerte!$L$2:$L${MAX_OFFER_ROWS},0)),0))'
        ws[f"G{row}"] = f'=IF($B{row}="","Nessuna offerta ammessa","")'
        ws[f"H{row}"] = "Singolo"
    style_cells(ws, "A6:H9", OUTPUT)

    section_header(ws, 12, "Migliori combinatorie per coppia", "H")
    combo_headers = ["Coppia", "BidderId", "BidderNome", "Totale", "Ammissibile", "Ribasso", "Note", "Origine"]
    for col, header in enumerate(combo_headers, start=1):
        ws.cell(row=13, column=col).value = header
    style_header_row(ws, 13, len(combo_headers))
    for row, pair_id in enumerate(["L1+L2", "L2+L3", "L3+L4", "L1+L4"], start=14):
        ws[f"A{row}"] = pair_id
        ws[f"D{row}"] = f'=SUMPRODUCT(MAX((Combinatorie!$C$5:$C${MAX_COMBO_ROWS}=$A{row})*(Combinatorie!$N$5:$N${MAX_COMBO_ROWS}="SI")*Combinatorie!$Q$5:$Q${MAX_COMBO_ROWS}))'
        ws[f"B{row}"] = f'=IF($D{row}=0,"",IFERROR(INDEX(Combinatorie!$A$5:$A${MAX_COMBO_ROWS},MATCH($A{row}&"|"&TEXT($D{row},"0.0000"),Combinatorie!$V$5:$V${MAX_COMBO_ROWS},0)),""))'
        ws[f"C{row}"] = f'=IF($B{row}="","",IFERROR(INDEX(Combinatorie!$B$5:$B${MAX_COMBO_ROWS},MATCH($A{row}&"|"&$B{row},Combinatorie!$W$5:$W${MAX_COMBO_ROWS},0)),""))'
        ws[f"E{row}"] = f'=IF($B{row}="","NO","SI")'
        ws[f"F{row}"] = f'=IF($B{row}="","",IFERROR(INDEX(Combinatorie!$E$5:$E${MAX_COMBO_ROWS},MATCH($A{row}&"|"&$B{row},Combinatorie!$W$5:$W${MAX_COMBO_ROWS},0)),0))'
        ws[f"G{row}"] = f'=IF($B{row}="","Nessuna combinatoria ammessa","")'
        ws[f"H{row}"] = "Combinatoria"
    style_cells(ws, "A14:H17", OUTPUT)

    section_header(ws, 20, "Matrice scenario indicativa", "H")
    scenario_headers = ["Scenario", "L1", "L2", "L3", "L4", "Totale", "Nota", "Rank"]
    for col, header in enumerate(scenario_headers, start=1):
        ws.cell(row=21, column=col).value = header
    style_header_row(ws, 21, len(scenario_headers))
    scenarios = [
        ("Tutti singoli", "Singolo L1", "Singolo L2", "Singolo L3", "Singolo L4", "SUM($D$6:$D$9)"),
        ("L1+L2 + singoli", "Combo L1+L2", "Combo L1+L2", "Singolo L3", "Singolo L4", "$D$14+$D$8+$D$9"),
        ("L2+L3 + singoli", "Singolo L1", "Combo L2+L3", "Combo L2+L3", "Singolo L4", "$D$6+$D$15+$D$9"),
        ("L3+L4 + singoli", "Singolo L1", "Singolo L2", "Combo L3+L4", "Combo L3+L4", "$D$6+$D$7+$D$16"),
        ("L1+L4 + singoli", "Combo L1+L4", "Singolo L2", "Singolo L3", "Combo L1+L4", "$D$17+$D$7+$D$8"),
        ("L1+L2 e L3+L4", "Combo L1+L2", "Combo L1+L2", "Combo L3+L4", "Combo L3+L4", "$D$14+$D$16"),
        ("L2+L3 e L1+L4", "Combo L1+L4", "Combo L2+L3", "Combo L2+L3", "Combo L1+L4", "$D$15+$D$17"),
    ]
    for row, scenario in enumerate(scenarios, start=22):
        name, l1, l2, l3, l4, total_formula = scenario
        ws[f"A{row}"] = name
        ws[f"B{row}"] = l1
        ws[f"C{row}"] = l2
        ws[f"D{row}"] = l3
        ws[f"E{row}"] = l4
        ws[f"F{row}"] = "=" + total_formula
        ws[f"G{row}"] = f'=IF($F{row}=0,"Nessun dato sufficiente",IF($H{row}=1,"Scenario migliore nella matrice",""))'
        ws[f"H{row}"] = f'=IF($F{row}=0,"",RANK($F{row},$F$22:$F$28,0))'
    style_cells(ws, "A22:H28", WHITE)
    ws.conditional_formatting.add("H22:H28", CellIsRule(operator="equal", formula=["1"], fill=fill(GREEN)))

    section_header(ws, 31, "Limiti della matrice Excel", "H")
    notes = [
        "La matrice è pensata per analisi operative rapide: considera singoli e coppie compatibili principali.",
        "Per warning documentali avanzati, persistenza JSON, simulazioni batch e deroga al limite di due lotti resta più affidabile la web app.",
        "Se una combinatoria non compare, controlla foglio Combinatorie: singoli attivi, soglia, buste, PEF e ribasso migliorativo.",
    ]
    for row, text in enumerate(notes, start=32):
        ws[f"A{row}"] = text
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"].fill = fill(AMBER if row == 32 else WHITE)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"A{row}"].border = BORDER

    set_widths(ws, {"A": 24, "B": 18, "C": 18, "D": 18, "E": 18, "F": 14, "G": 34, "H": 10})
    ws.freeze_panes = "A5"


def excel_json_text(ref: str) -> str:
    return f'SUBSTITUTE({ref},CHAR(34),"{chr(39)}")'


def excel_number_value(ref: str) -> str:
    return f'IFERROR({ref}*1,IFERROR(NUMBERVALUE({ref},".",","),IFERROR(NUMBERVALUE({ref},",","."),0)))'


def excel_json_number(ref: str) -> str:
    return f'SUBSTITUTE(TRIM(ROUND({excel_number_value(ref)},4)&""),",",".")'


def excel_json_boolean(ref: str) -> str:
    return f'IF({excel_number_value(ref)}=1,"true","false")'


def excel_timestamp() -> str:
    return 'YEAR(NOW())&TEXT(MONTH(NOW()),"00")&TEXT(DAY(NOW()),"00")&TEXT(HOUR(NOW()),"00")&TEXT(MINUTE(NOW()),"00")&TEXT(SECOND(NOW()),"00")'


def excel_iso_date() -> str:
    return 'YEAR(TODAY())&"-"&TEXT(MONTH(TODAY()),"00")&"-"&TEXT(DAY(TODAY()),"00")'


def create_scambio_web(wb):
    ws = reset_sheet(wb, "ScambioWeb")
    apply_sheet_chrome(ws, BLUE)
    title(
        ws,
        "Scambio web",
        "JSON copiabile fra Excel e web: include offerte, sub-criteri A-G, ribassi e combinatorie.",
    )

    section_header(ws, 4, "Come usarlo", "H")
    steps = [
        ("1", "Compila Offerte e, se servono, Combinatorie."),
        ("2", "Esegui CheckBeforeRun e SimulaScenario per aggiornare formule e warning."),
        ("3", "Copia le righe non vuote della colonna A dalla sezione JSON generato."),
        ("4", "Salvale come file .json oppure incollale in un editor e importale dal simulatore web."),
    ]
    for row, (step, text) in enumerate(steps, start=5):
        ws[f"A{row}"] = step
        ws[f"B{row}"] = text
        ws.merge_cells(f"B{row}:H{row}")
    style_cells(ws, "A5:H8", WHITE)
    for row in range(5, 9):
        ws[f"A{row}"].fill = fill(NAVY)
        ws[f"A{row}"].font = Font(color=WHITE, bold=True)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")

    section_header(ws, 10, "JSON generato", "H")
    ws["A11"] = "Copia la colonna A. Le righe vuote sono solo spazi bianchi JSON e possono essere ignorate."
    ws.merge_cells("A11:H11")
    ws["A11"].fill = fill(AMBER)
    ws["A11"].border = BORDER
    ws["A11"].alignment = Alignment(wrap_text=True)

    row = 13

    def put(value):
        nonlocal row
        ws[f"A{row}"] = value
        row += 1

    put("{")
    put('  "format": "glm-excel-v1",')
    put('  "schemaVersion": 1,')
    put('="  "&CHAR(34)&"id"&CHAR(34)&": "&CHAR(34)&"excel-"&' + excel_timestamp() + '&CHAR(34)&","')
    put('="  "&CHAR(34)&"name"&CHAR(34)&": "&CHAR(34)&"Scenario Excel "&' + excel_iso_date() + '&CHAR(34)&","')
    put('  "baseScenarioId": "market",')
    put(
        '="  "&CHAR(34)&"settings"&CHAR(34)&": "&CHAR(123)&'
        'CHAR(34)&"threshold"&CHAR(34)&": "&'
        + excel_json_number("Parametri!$B$2")
        + '&", "&CHAR(34)&"applyAwardLimitDerogation"&CHAR(34)&": false"&CHAR(125)&","'
    )
    put('="  "&CHAR(34)&"selectedLotId"&CHAR(34)&": "&CHAR(34)&Parametri!$B$3&CHAR(34)&","')
    put('  "selectedPairId": "L1+L2",')
    put('  "notes": "Export da Excel: offerte, sub-criteri A-G, ribassi e combinatorie.",')
    put('  "offers": [')

    for offer_row in range(2, MAX_OFFER_ROWS + 1):
        has_more = f'COUNTA(Offerte!$A{offer_row + 1}:$A${MAX_OFFER_ROWS})>0' if offer_row < MAX_OFFER_ROWS else "FALSE"
        bidder_id = excel_json_text(f"Offerte!$A{offer_row}")
        bidder_name = excel_json_text(f"Offerte!$B{offer_row}")
        lot_id = excel_json_text(f"Offerte!$C{offer_row}")
        enabled = excel_json_boolean(f"Offerte!$D{offer_row}")
        technical = excel_json_number(f"Offerte!$E{offer_row}")
        discount = excel_json_number(f"Offerte!$F{offer_row}")
        put(
            f'=IF(Offerte!$A{offer_row}="","",'
            f'"    "&CHAR(123)&CHAR(34)&"bidderId"&CHAR(34)&": "&CHAR(34)&{bidder_id}&CHAR(34)&", "&'
            f'CHAR(34)&"bidderName"&CHAR(34)&": "&CHAR(34)&{bidder_name}&CHAR(34)&", "&'
            f'CHAR(34)&"lotId"&CHAR(34)&": "&CHAR(34)&{lot_id}&CHAR(34)&", "&'
            f'CHAR(34)&"enabled"&CHAR(34)&": "&{enabled}&", "&'
            f'CHAR(34)&"technicalRaw"&CHAR(34)&": "&{technical}&", "&'
            f'CHAR(34)&"discount"&CHAR(34)&": "&{discount}&CHAR(125)'
            f'&IF({has_more},",",""))'
        )

    put("  ],")
    put('  "criteria": [')
    criteria_count = len(load_criteria())
    criteria_rows_end = TECHNICAL_START_ROW + (MAX_OFFER_ROWS - 1) * criteria_count - 1
    for technical_row in range(TECHNICAL_START_ROW, criteria_rows_end + 1):
        criterion_index = (technical_row - TECHNICAL_START_ROW) % criteria_count
        offer_row = 2 + (technical_row - TECHNICAL_START_ROW) // criteria_count
        has_more = "TRUE" if criterion_index < criteria_count - 1 else f'COUNTA(Offerte!$A{offer_row + 1}:$A${MAX_OFFER_ROWS})>0'
        bidder_id = excel_json_text(f"CriteriTecnici!$A{technical_row}")
        lot_id = excel_json_text(f"CriteriTecnici!$C{technical_row}")
        criterion_id = excel_json_text(f"CriteriTecnici!$D{technical_row}")
        kind = excel_json_text(f"CriteriTecnici!$F{technical_row}")
        value = excel_json_number(f"CriteriTecnici!$N{technical_row}")
        numerator = excel_json_number(f"CriteriTecnici!$K{technical_row}")
        denominator = excel_json_number(f"CriteriTecnici!$L{technical_row}")
        flag = excel_json_number(f"CriteriTecnici!$M{technical_row}")
        raw_score = excel_json_number(f"CriteriTecnici!$O{technical_row}")
        put(
            f'=IF(CriteriTecnici!$A{technical_row}="","",'
            f'"    "&CHAR(123)&CHAR(34)&"bidderId"&CHAR(34)&": "&CHAR(34)&{bidder_id}&CHAR(34)&", "&'
            f'CHAR(34)&"lotId"&CHAR(34)&": "&CHAR(34)&{lot_id}&CHAR(34)&", "&'
            f'CHAR(34)&"criterionId"&CHAR(34)&": "&CHAR(34)&{criterion_id}&CHAR(34)&", "&'
            f'CHAR(34)&"kind"&CHAR(34)&": "&CHAR(34)&{kind}&CHAR(34)&", "&'
            f'CHAR(34)&"value"&CHAR(34)&": "&{value}&", "&'
            f'CHAR(34)&"numerator"&CHAR(34)&": "&{numerator}&", "&'
            f'CHAR(34)&"denominator"&CHAR(34)&": "&{denominator}&", "&'
            f'CHAR(34)&"flag"&CHAR(34)&": "&{flag}&", "&'
            f'CHAR(34)&"rawScore"&CHAR(34)&": "&{raw_score}&CHAR(125)'
            f'&IF({has_more},",",""))'
        )
    put("  ],")
    put('  "combos": [')
    for combo_row in range(5, MAX_COMBO_ROWS + 1):
        has_more = f'COUNTA(Combinatorie!$A{combo_row + 1}:$A${MAX_COMBO_ROWS})>0' if combo_row < MAX_COMBO_ROWS else "FALSE"
        bidder_id = excel_json_text(f"Combinatorie!$A{combo_row}")
        bidder_name = excel_json_text(f"Combinatorie!$B{combo_row}")
        pair_id = excel_json_text(f"Combinatorie!$C{combo_row}")
        enabled = excel_json_boolean(f"Combinatorie!$D{combo_row}")
        discount = excel_json_number(f"Combinatorie!$E{combo_row}")
        inserted = excel_json_boolean(f"Combinatorie!$F{combo_row}")
        pef = excel_json_boolean(f"Combinatorie!$G{combo_row}")
        put(
            f'=IF(Combinatorie!$A{combo_row}="","",'
            f'"    "&CHAR(123)&CHAR(34)&"bidderId"&CHAR(34)&": "&CHAR(34)&{bidder_id}&CHAR(34)&", "&'
            f'CHAR(34)&"bidderName"&CHAR(34)&": "&CHAR(34)&{bidder_name}&CHAR(34)&", "&'
            f'CHAR(34)&"pairId"&CHAR(34)&": "&CHAR(34)&{pair_id}&CHAR(34)&", "&'
            f'CHAR(34)&"enabled"&CHAR(34)&": "&{enabled}&", "&'
            f'CHAR(34)&"discount"&CHAR(34)&": "&{discount}&", "&'
            f'CHAR(34)&"insertedInBothBuste"&CHAR(34)&": "&{inserted}&", "&'
            f'CHAR(34)&"pefCoherent"&CHAR(34)&": "&{pef}&CHAR(125)'
            f'&IF({has_more},",",""))'
        )
    put("  ]")
    put("}")

    style_cells(ws, f"A13:A{row - 1}", WHITE)
    for json_row in range(13, row):
        ws[f"A{json_row}"].font = Font(name="Menlo", size=9, color=INK)
        ws[f"A{json_row}"].alignment = Alignment(wrap_text=False, vertical="top")
    ws.column_dimensions["A"].width = 145
    set_widths(ws, {"B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12, "H": 12})
    ws.freeze_panes = "A13"


def polish_results(wb):
    ws = wb["Risultati"]
    clear_layout_helpers(ws)
    apply_sheet_chrome(ws, TEAL_DARK)
    if ws.max_row >= 1:
        style_header_row(ws, 1, min(ws.max_column, 8))
    style_cells(ws, f"A2:H{max(ws.max_row, 30)}", OUTPUT)
    ws.conditional_formatting.add("D2:D200", CellIsRule(operator="equal", formula=['"NO"'], fill=fill(RED)))
    ws.conditional_formatting.add("D2:D200", CellIsRule(operator="equal", formula=['"SI"'], fill=fill(GREEN)))
    set_widths(ws, {"A": 18, "B": 16, "C": 24, "D": 14, "E": 14, "F": 14, "G": 14, "H": 30})
    ws.freeze_panes = "A2"


def polish_confronto(wb):
    ws = wb["ConfrontoWeb"]
    clear_layout_helpers(ws)
    apply_sheet_chrome(ws, BLUE)
    ws["I1"] = "Input expected"
    ws["J1"] = "Valore"
    style_header_row(ws, 1, 10)
    expected_labels = ["Web L1", "Web L2", "Web L3", "Web L4"]
    for idx, label in enumerate(expected_labels, start=2):
        ws[f"I{idx}"] = label
        if ws[f"J{idx}"].value is None:
            ws[f"J{idx}"] = 0
        ws[f"J{idx}"].fill = fill(INPUT)
    style_cells(ws, f"A2:J{max(ws.max_row, 8)}", WHITE)
    dv_expected = DataValidation(type="decimal", operator="between", formula1="0", formula2="100")
    ws.add_data_validation(dv_expected)
    dv_expected.add("J2:J5")
    add_table(ws, "tblConfrontoWeb", f"A1:J{max(ws.max_row, 8)}")
    set_widths(ws, {"A": 18, "B": 22, "C": 12, "D": 16, "E": 12, "F": 34, "G": 14, "H": 34, "I": 18, "J": 14})
    ws.freeze_panes = "A2"


def polish_log(wb):
    ws = wb["LogOttimizzazione"]
    clear_layout_helpers(ws)
    apply_sheet_chrome(ws, TEAL_DARK)
    style_header_row(ws, 1, max(ws.max_column, 7))
    style_cells(ws, f"A2:J{max(ws.max_row, 50)}", OUTPUT)
    set_widths(ws, {"A": 14, "B": 16, "C": 12, "D": 12, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16, "J": 20})
    ws.freeze_panes = "A2"


def add_common_footer(wb):
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0


def main():
    path = workbook_path()
    wb = load_workbook(path, keep_vba=True)

    create_dashboard(wb)
    create_guide(wb)
    create_glossary(wb)
    polish_instruction_sheet(wb)
    polish_parametri(wb)
    polish_ottimizzazione(wb)
    polish_offerte(wb)
    create_criteri_tecnici(wb)
    create_combinatorie(wb)
    create_scenario_globale(wb)
    create_scambio_web(wb)
    polish_results(wb)
    polish_confronto(wb)
    polish_log(wb)
    add_common_footer(wb)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    reorder_sheets(
        wb,
        [
            "Dashboard",
            "Parametri",
            "CriteriTecnici",
            "Offerte",
            "Ottimizzazione",
            "Combinatorie",
            "Risultati",
            "Guida",
            "Istruzioni",
            "ScenarioGlobale",
            "ScambioWeb",
            "ConfrontoWeb",
            "LogOttimizzazione",
            "Glossario",
        ],
    )
    apply_workbook_visibility(wb)

    for ws in wb.worksheets:
        ws.sheet_view.tabSelected = False
    wb["Dashboard"].sheet_view.tabSelected = True
    wb.active = wb.sheetnames.index("Dashboard")
    wb.save(path)
    print(f"Workbook Excel migliorato: {path}")


if __name__ == "__main__":
    main()
